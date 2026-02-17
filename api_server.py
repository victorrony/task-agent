"""
API SERVER - FinanceAgent Pro REST API
=====================================
Backend industrial para o frontend Next.js.
Exclui: Gradio (substituído por esta API)
Inclui: Endpoints RESTful, CORS, Pydantic Models.
"""

from fastapi import FastAPI, HTTPException, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
from agent import TaskAgent, init_db
from agent.data_service import get_quick_stats, get_transactions_df, get_goals_progress, get_users, get_expense_categories
from agent.tools import set_user_id

# Inicialização
init_db()
app = FastAPI(title="FinanceAgent Pro API", version="3.1.0")

# CORS (Permitir Frontend Next.js)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Em produção, restringir para o domínio do frontend
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- MODELOS DE DADOS (Schemas) ---

class ChatRequest(BaseModel):
    user_id: int
    message: str
    mode: str = "assistant"

class Transaction(BaseModel):
    date: str
    description: str
    amount: float
    type: str
    category: str

# --- PROCESSAMENTO DE FICHEIROS ---

TEXT_EXTENSIONS = {".txt", ".md", ".json", ".xml", ".html", ".htm", ".log", ".css", ".js", ".py", ".csv"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}

def _extract_file_content(filename: str, content_bytes: bytes) -> str:
    """Extrai texto de vários tipos de ficheiro."""
    import io, os
    ext = os.path.splitext(filename)[1].lower()

    # PDF
    if ext == ".pdf":
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(content_bytes))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except Exception as e:
            return f"[Erro ao ler PDF: {e}]"

    # Word (.docx)
    if ext == ".docx":
        try:
            from docx import Document
            doc = Document(io.BytesIO(content_bytes))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            return f"[Erro ao ler Word: {e}]"

    # Excel (.xlsx, .xls)
    if ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content_bytes), read_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"--- Folha: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    lines.append(" | ".join(str(cell) if cell is not None else "" for cell in row))
            return "\n".join(lines)
        except Exception as e:
            return f"[Erro ao ler Excel: {e}]"

    # CSV
    if ext == ".csv":
        try:
            import csv
            text = content_bytes.decode("utf-8")
            reader = csv.reader(io.StringIO(text))
            return "\n".join(" | ".join(row) for row in reader)
        except Exception as e:
            return f"[Erro ao ler CSV: {e}]"

    # Imagens - converter para base64 para o Gemini interpretar via visão
    if ext in IMAGE_EXTENSIONS:
        import base64
        b64 = base64.b64encode(content_bytes).decode("utf-8")
        mime = f"image/{'jpeg' if ext in ('.jpg', '.jpeg') else ext.lstrip('.')}"
        return f"[IMAGEM base64 mime={mime}]\n{b64}"

    # Texto genérico
    if ext in TEXT_EXTENSIONS:
        try:
            return content_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return content_bytes.decode("latin-1", errors="replace")

    # Fallback: tenta decodificar como texto
    try:
        return content_bytes.decode("utf-8")
    except Exception:
        return "[Formato de ficheiro não suportado para extração de texto]"


# --- ENDPOINTS ---

@app.get("/")
def read_root():
    return {"status": "online", "system": "FinanceAgent Pro v3.1"}

@app.get("/users")
def list_users():
    """Listar utilizadores disponíveis para login."""
    return get_users()

@app.get("/dashboard/{user_id}")
def get_dashboard_data(user_id: int):
    """Retorna estatísticas consolidadas para o dashboard."""
    set_user_id(user_id)
    stats = get_quick_stats(user_id)
    goals = get_goals_progress(user_id)
    # Transactions (últimas 5 para preview)
    df = get_transactions_df(user_id)
    recent_tx = df.head(5).to_dict(orient="records") if not df.empty else []
    
    return {
        "stats": stats,
        "goals": goals,
        "recent_transactions": recent_tx
    }

def _detect_action(response_text: str) -> str | None:
    """Detect what action the agent performed based on response patterns."""
    text = response_text.lower()
    if any(w in text for w in ['saldo definido', 'saldo atualizado', 'balance set', 'balance updated']):
        return 'balance_updated'
    if any(w in text for w in ['transação registada', 'transação adicionada', 'transaction added', 'transaction recorded', 'registado com sucesso', 'despesa registada', 'receita registada']):
        return 'transaction_added'
    if any(w in text for w in ['meta criada', 'meta atualizada', 'goal created', 'goal updated']):
        return 'goal_updated'
    if any(w in text for w in ['preferência', 'perfil atualizado', 'preference saved', 'profile updated']):
        return 'preference_saved'
    return None


def _process_agent_response(user_id: int, message: str, mode: str) -> dict:
    """Executa o agente e normaliza a resposta."""
    import re
    set_user_id(user_id)
    agent = TaskAgent(user_id=user_id, mode=mode)
    response = agent.run(message)

    # Normalização robusta para Gemini 2.5
    if isinstance(response, list):
        text_parts = []
        for item in response:
            if isinstance(item, dict):
                text_parts.append(item.get("text", ""))
            elif isinstance(item, str):
                text_parts.append(item)
            else:
                text_parts.append(str(item))
        response = "".join(text_parts)
    elif isinstance(response, dict):
        response = response.get("text", response.get("content", str(response)))
    elif not isinstance(response, str):
        response = str(response) if response else ""

    # Remove metadata do Gemini
    response = re.sub(r"['\"]extras['\"]\s*:\s*\{[^}]*\}", "", response)
    response = re.sub(r"['\"]signature['\"]\s*:\s*['\"][^'\"]*['\"]", "", response)
    response = re.sub(r"['\"]type['\"]\s*:\s*['\"]text['\"]", "", response)
    response = re.sub(r",\s*}", "}", response)
    response = re.sub(r"{\s*,", "{", response)
    response = re.sub(r"{\s*}", "", response)

    # Limpa espaços duplicados preservando \n
    lines = response.split("\n")
    lines = [re.sub(r" {2,}", " ", line).strip() for line in lines]
    response = "\n".join(lines).strip()

    return {"response": response, "action": _detect_action(response)}


@app.post("/chat", status_code=200)
async def chat_agent(
    user_id: int = Form(None),
    message: str = Form(None),
    mode: str = Form("assistant"),
    file: Optional[UploadFile] = File(None),
):
    """Endpoint de conversa com ficheiros (multipart/form-data)."""
    try:
        if user_id is None:
            raise HTTPException(status_code=400, detail="user_id is required")

        final_message = message or ""

        # Processamento de ficheiros (multi-formato)
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        if file:
            content_bytes = await file.read()
            if len(content_bytes) > MAX_FILE_SIZE:
                raise HTTPException(status_code=413, detail="Ficheiro excede o limite de 10MB")

            filename = file.filename.lower()
            file_content = _extract_file_content(filename, content_bytes)
            final_message = f"{final_message}\n\n[CONTEÚDO DO DOCUMENTO: {file.filename}]\n{file_content}"

        return _process_agent_response(user_id, final_message, mode)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/chat/json", status_code=200)
async def chat_agent_json(req: ChatRequest):
    """Endpoint de conversa via JSON (sem ficheiros)."""
    try:
        set_user_id(req.user_id)
        return _process_agent_response(req.user_id, req.message, req.mode)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/chat/history/{user_id}")
def get_chat_history(user_id: int):
    """Retorna o histórico de mensagens da sessão atual."""
    try:
        import re as re_hist
        agent = TaskAgent(user_id=user_id)
        history = agent.memory.get_history(limit=50)

        clean_history = []
        for msg in history:
            role = "unknown"
            if msg.type == "human": role = "user"
            elif msg.type == "ai": role = "assistant"
            elif msg.type == "system": continue
            elif msg.type == "tool": continue

            content = msg.content
            # Normalizar conteudo (Gemini 2.5 pode ter salvo listas)
            if isinstance(content, list):
                content = "".join(
                    item.get("text", "") if isinstance(item, dict) else str(item)
                    for item in content
                )
            else:
                content = str(content)

            # Limpar metadata residual
            content = re_hist.sub(r"['\"]extras['\"]\s*:\s*\{[^}]*\}", "", content)
            content = re_hist.sub(r"['\"]signature['\"]\s*:\s*['\"][^'\"]*['\"]", "", content)
            content = re_hist.sub(r"['\"]type['\"]\s*:\s*['\"]text['\"]", "", content)
            content = re_hist.sub(r",\s*}", "}", content)
            content = re_hist.sub(r"{\s*,", "{", content)
            content = re_hist.sub(r"{\s*}", "", content)
            content = content.strip()

            if content:
                clean_history.append({"role": role, "content": content})

        return clean_history
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/expenses/categories/{user_id}")
def get_expense_categories_endpoint(user_id: int):
    """Distribuição de gastos por categoria (últimos 90 dias)."""
    return get_expense_categories(user_id)

@app.get("/transactions/{user_id}")
def get_transactions(user_id: int):
    """Histórico completo de transações."""
    df = get_transactions_df(user_id)
    return df.to_dict(orient="records") if not df.empty else []

if __name__ == "__main__":
    import uvicorn
    # Usa porta 8005 (menos conflito)
    uvicorn.run(app, host="0.0.0.0", port=8005)
